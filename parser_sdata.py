
import sys

from PyQt5.QtWidgets import QFileDialog
from PyQt5 import QtWidgets

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ui.ui_main import Ui_MainWindow
from bots.google_bot import GoogleBot
from bots.yandex_bot import YandexBot
from models.url_model import UrlModel


class main_window(QtWidgets.QMainWindow, Ui_MainWindow):

    def __init__(self):
        super(main_window, self).__init__()
        self.setupUi(self)
        self.file_change.clicked.connect(self.openFileNameDialog)
        self.push_go.clicked.connect(self.startParser)
        self.setWindowTitle('Приложение')

        self.sheet_cache = object()


    def startParser(self):
        url_bd = UrlModel('test')

        url_book = self.file_url.text()
        book = load_workbook(url_book)
        self.sheet_cache = book.active
        for row in range(self.sheet_cache.max_row):
            if self.sheet_cache[f'B{row+2}'].value == None:
                break
            url_bd.load_url(self.sheet_cache[f'B{row+2}'].value)

        google_bot = GoogleBot(url_bd)
        google_bot.iter_urls()
        result_google = url_bd.select_result(
            'data_google', 
            'current_google'
            )

        yandex_bot = YandexBot(url_bd)
        yandex_bot.iter_urls()
        result_yandex = url_bd.select_result(
            'data_yandex', 
            'current_yandex'
            )

        self.__fillResult(result_google, 'D')
        self.__fillResult(result_yandex, 'E')
        
        book.save('test.xlsx')


    def __fillResult(self, result, str_column=None):
        true_fill = PatternFill('solid', fgColor='0000FF00')
        false_fill = PatternFill('solid', fgColor='00FF0000')
        for id, row in enumerate(range(self.sheet_cache.max_row-1)):
            try:
                self.sheet_cache[f'{str_column}{row+2}'].value = result[id][0]
                if result[id][1]:
                    self.sheet_cache[f'{str_column}{row+2}'].fill = true_fill
                else:
                    self.sheet_cache[f'{str_column}{row+2}'].fill = false_fill
            except:
                pass
            

    #Один файл
    def openFileNameDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","All Files (*);;Python Files (*.py)", options=options)
        if fileName:
            self.file_url.setText(fileName)
    

    def openFileNamesDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        files, _ = QFileDialog.getOpenFileNames(self,"QFileDialog.getOpenFileNames()", "","All Files (*);;Python Files (*.py)", options=options)
        if files:
            print(files)
    

    def saveFileDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","All Files (*);;Text Files (*.txt)", options=options)
        if fileName:
            print(fileName)

    
app = QtWidgets.QApplication([])
application = main_window()
application.show()

sys.exit(app.exec())

